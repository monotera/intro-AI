import pandas as pd
from openpyxl import load_workbook
import pprint
import json
import itertools


def get_data(filepath):
    # Obtains sheet names from excel
    sheetnames = get_sheetnames_xlsx(filepath)
    prob_data = {}
    """
    - Obtains the data from each sheet in excel.
    - it comes as a dataFrame and then its transform into a dictionary
    example of data:
    prob_data =
    {   'Appointment': {'Train': ['on time', 'delayed'],
                        'attend': [0.9, 0.6],
                        'miss': [0.1, 0.4]},
        'Maintenance': {'Rain': ['none', 'light', 'heavy'],
                        'no': [0.6, 0.8, 0.9],
                        'yes': [0.4, 0.2, 0.1]},
        'Rain':        {'heavy': [0.1], 'light': [0.2], 'none': [0.7]},
        'Train':       {'Maintenance': ['yes', 'no', 'yes', 'no', 'yes', 'no'],
                        'Rain': ['none', 'none', 'light', 'light', 'heavy', 'heavy'],
                        'delayed': [0.2, 0.1, 0.4, 0.3, 0.6, 0.5],
                        'on time': [0.8, 0.9, 0.6, 0.7, 0.4, 0.5]}
    }
    """
    for sheetname in sheetnames:
        prob_dataFrame = (pd.read_excel(filepath, sheet_name=sheetname))
        prob_data[sheetname] = prob_dataFrame.to_dict()

    # Transforms inside dictionary into a list
    for key, value in prob_data.items():
        for k, v in value.items():
            value[k] = list(v.values())
    return sheetnames, prob_data


def get_sheetnames_xlsx(filepath):
    wb = load_workbook(filepath, read_only=True, keep_links=False)
    return wb.sheetnames


"""
the syntaxis of the query basic should be:
query_var   = The variable for which you want to calculate the probability distribution
query_data  = one or more observed variables for an event e
json example:
{
    "query_var": "Appointment",
    "query_data": {
        "Rain": "light",
        "Maintenance": "no",
    }
}
"""


def get_basic_query(filepath):
    with open(filepath) as json_file:
        query = json.load(json_file)
    return query


"""
Completes the query with the unknown variables and the format needen for the script
query_var   = The variable for which you want to calculate the probability distribution
query_data  = one or more observed variables for an event e, if not known then the event is "?"
unknown_var = variables that are not the query and have not been observed
json example:
{
    "query_var": "Appointment",
    "query_data": {
        "Rain": "light",
        "Maintenance": "no",
        "Train": "?",
        "Appointment": "?"
    }
    unknown_var : ["Train]
}
"""


def build_complete_query(data, query, data_names):
    query_data = query["query_data"]
    query_data[query["query_var"]] = "?"
    unknown_var = []
    vars_needed = get_var_needed(query, data_names, data)
    for var_needed in vars_needed:
        if var_needed not in query_data:
            query_data[var_needed] = "?"
            unknown_var.append(var_needed)
    query["unknown_var"] = unknown_var
    return query


"""
obtains the variables needed for resolving the probability of 
the variable for which you want to calculate the probability distribution
example: for Train it would be [Train,Rain,Maintenance]
"""


def get_var_needed(query, data_names, data):
    vars_needed = []
    dependencies = [query["query_var"]]
    while len(dependencies) > 0:
        aux = dependencies[0]
        del dependencies[0]
        if aux not in vars_needed:
            vars_needed.append(aux)
        dependencies += find_dependencies_of_var(data[aux], data_names)

    return vars_needed


"""
gets the options of certain variable,
example: the options for Rain are [heavy, light, none]
"""


def get_var_options(data, data_names):
    options = []
    for k in data:
        if k not in data_names:
            options.append(k)
    return options


"""
Gets all the combinations of the options of the unknown variables.
For example if the unknown variables are Rain and Train the combinations
would be:
[('Train:on time', 'Rain:none'), ('Train:on time', 'Rain:light'),
('Train:on time', 'Rain:heavy'), ('Train:delayed', 'Rain:none'),
('Train:delayed', 'Rain:light'), ('Train:delayed', 'Rain:heavy')]
"""


def get_combinations_unknown_vars(unknown_vars, data_names, data):
    unknown_vars_options = []
    for var in unknown_vars:
        var_options = get_var_options(data[var], data_names)
        var_options = [var + ":"+s for s in var_options]
        unknown_vars_options.append(var_options)
    combinations = list(itertools.product(*unknown_vars_options))
    return combinations


def get_formulas(query, query_var_options):
    formulas = []
    for option in query_var_options:
        formula = "P(" + query["query_var"]+":" + option + " | "
        for k, v in query["query_data"].items():
            if v != "?" and v:
                formula += k + ":" + v + " ^ "
        formula = formula[:-2]
        formula += ") = "
        formulas.append(formula)
    return formulas


"""
Fills the missing data in query_data depending on the combination
"""


def fill_query_data_with_com(combination, query_data):
    for com in combination:
        com_part = com.partition(":")
        query_data[com_part[0]] = com_part[2]
    return query_data


"""
Calculates the probability of a query,
if a variable of the query doesnt depends on another:
        -it returns the value depending on the query,
if the variable depends on another variable:
        -finds the corresponding index and then it returns the value depending on the query and the index
after finding the probability of a variables, its multiply to the rest of probabilities
returns the multiplication of all query variables probabilities

example incoming data:
    query_data = {'Rain': 'light', 'Maintenance': 'no',
        'Appointment': 'attend', 'Train': 'on time'}
    data =
        {   'Appointment': {'Train': ['on time', 'delayed'],
                            'attend': [0.9, 0.6],
                            'miss': [0.1, 0.4]},
            'Maintenance': {'Rain': ['none', 'light', 'heavy'],
                            'no': [0.6, 0.8, 0.9],
                            'yes': [0.4, 0.2, 0.1]},
            'Rain':        {'heavy': [0.1], 'light': [0.2], 'none': [0.7]},
            'Train':       {'Maintenance': ['yes', 'no', 'yes', 'no', 'yes', 'no'],
                            'Rain': ['none', 'none', 'light', 'light', 'heavy', 'heavy'],
                            'delayed': [0.2, 0.1, 0.4, 0.3, 0.6, 0.5],
                            'on time': [0.8, 0.9, 0.6, 0.7, 0.4, 0.5]}
        }
    data_names = [Rain, Appointment, Train, Main]
example return : 0.10080
"""


def find_probability(data, query_data, data_names):
    final_prob = 1
    local_prob = 0
    print_query(data, query_data, data_names)
    for query_var in data_names:
        if query_var in query_data:
            dependencies = find_dependencies_of_var(
                data[query_var], data_names)
            option = query_data[query_var]
            probs = data[query_var][option]
            if len(dependencies) == 0:
                local_prob = probs[0]
            else:
                index = find_index(data[query_var], dependencies, query_data)
                local_prob = probs[index]
            final_prob *= local_prob

    return final_prob


"""
Finds the index where the value of this variables match the value of the query
for example:
'Train':       {'Maintenance': ['yes', 'no', 'yes', 'no', 'yes', 'no'],
                        'Rain': ['none', 'none', 'light', 'light', 'heavy', 'heavy'],
                        'delayed': [0.2, 0.1, 0.4, 0.3, 0.6, 0.5],
                        'on time': [0.8, 0.9, 0.6, 0.7, 0.4, 0.5]}
query_data = {'Rain': 'light', 'Maintenance': 'no',
    'Appointment': 'attend', 'Train': 'on time'}
Train depends on Maintenance and Rain, this variables in the query have the value of ['Maintenance:no','Rain:light'],
so the index would be 3
"""


def find_index(data, dependencies, query_data):
    dep_value = []
    dep_options = []
    for dep in dependencies:
        dep_value.append(dep+":"+query_data[dep])
        aux = [dep+":" + value for value in data[dep]]
        dep_options.append(aux)
    for i in range(len(data[dep])):
        find_aux = []
        for j in range(len(dependencies)):
            find_aux.append(dep_options[j][i])
        if find_aux == dep_value:
            return i
    return -1


"""
If a variable depends on others to find its probability,
returns the name of this variables
example:
'Maintenance': {'Rain': ['none', 'light', 'heavy'],
                        'no': [0.6, 0.8, 0.9],
                        'yes': [0.4, 0.2, 0.1]},
The dependency of Maintenance would be [Rain]
"""


def find_dependencies_of_var(data, data_names):
    dependencies = []
    for k, v in data.items():
        if k in data_names:
            dependencies.append(k)
    return dependencies


def print_query(data, query, data_names):
    for query_var in query:
        prefix = "P(" + query_var+":"+query[query_var]
        deps = find_dependencies_of_var(data[query_var], data_names)
        if len(deps) == 0:
            print(prefix+")", end=" ")
        else:
            sufix = " | " + \
                "".join([var+":"+query[var] + " ^ " for var in deps])
            sufix = sufix[:-2] + ")"
            print(prefix + sufix, end=" ")
    print(" = ", end=" ")


def transform_into_prob(probs, query_var_options):
    suma = sum(probs)
    alpha = 1/suma
    transform_probs = {}
    for i in range(len(probs)):
        transform_probs[query_var_options[i]] = probs[i]*alpha
    return transform_probs


def __main__():
    options_excel = {
        "clase": "./Datos-IA-3.xlsx",
        "Ejercicio1": "./universidad.xlsx",
        "Ejercicio2": "./felicidad.xlsx"
    }
    pprint.pprint(options_excel)
    excel_opc = input("Porfavor escriba una opcion (ej : Ejercicio1) : ")
    excel_filepath = options_excel[excel_opc]
    print(excel_filepath)

    options_json = {
        "clase": {
            "query1": "./testApp.json"
        },
        "Ejercicio1": {
            "query1": "./testUni1.json",
            "query2": "./testUni2.json"
        },
        "Ejercicio2": {
            "query1": "./testHappy1.json",
            "query2": "./testHappy2.json"
        }
    }
    pprint.pprint(options_json[excel_opc])
    json_opc = input("Porfavor escriba una opcion (ej : query1) : ")
    query_filepath = options_json[excel_opc][json_opc]

    # Obtener datos
    data_names, data = get_data(excel_filepath)
    query = get_basic_query(query_filepath)
    query = build_complete_query(data, query, data_names)

    # Obtener informacion var objetivo
    query_var_data = data[query["query_var"]]
    query_var_options = get_var_options(query_var_data, data_names)
    query_data = query["query_data"]

    # Obtener combinacion de variables desconocidas
    unknown_com = get_combinations_unknown_vars(
        query["unknown_var"], data_names, data)

    formulas = get_formulas(query, query_var_options)
    probs = []
    for i in range(len(query_var_options)):
        print(formulas[i])
        query_data[query["query_var"]] = query_var_options[i]
        final_prob = 0
        if len(unknown_com) == 0:
            local_prob = find_probability(data, query_data, data_names)
            print(local_prob)
            final_prob += local_prob
        else:
            for combination in unknown_com:
                query_data = fill_query_data_with_com(combination, query_data)
                local_prob = find_probability(data, query_data, data_names)
                print(local_prob)
                final_prob += local_prob
        probs.append(final_prob)
        print("Result = ", final_prob)
    print(probs)
    print("Probabilidades:")
    print(transform_into_prob(probs, query_var_options))


__main__()
